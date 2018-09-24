#You can change the default voice by editing the line below:
standardVoice = "Wavenet-A"


import argparse
import os
import shutil
import sys

from tqdm import tqdm
from tqdm._utils import _term_move_up
from google.cloud import texttospeech
from pydub import AudioSegment
from openpyxl import load_workbook
from colorama import init, Fore, Back
init(autoreset=True) #initialize colorama

prefix = _term_move_up()

def synthesize_ssml(ssml_text, output_file_name):
	"""Synthesizes speech from the input text of ssml.
	Note: documentation from google for ssml
		https://cloud.google.com/text-to-speech/docs/ssml
	"""

	input_text = texttospeech.types.SynthesisInput(ssml=ssml_text)

	audio_config = texttospeech.types.AudioConfig(
        audio_encoding=texttospeech.enums.AudioEncoding.MP3)
	
	
	if os.path.isfile(output_file_name + ".wav") and (not args.override):
		tqdm.write(prefix + Fore.BLUE + "Skipped -->")
	else:
		if os.path.isfile(output_file_name + ".wav"):
			tqdm.write(prefix + Fore.GREEN + "Overwritten -->")
		response = client.synthesize_speech(input_text, voice, audio_config)
		# The response's audio_content is binary.
		with open('output.mp3', 'wb') as out:
			out.write(response.audio_content)
		sound = AudioSegment.from_mp3("output.mp3")
		sound = sound.set_channels(1)
		sound = sound.set_frame_rate(32000)
		sound.export(output_file_name + ".wav", format="wav")


def create_filesystem(filepath):
	if args.remove and os.path.exists(filepath):
		print(Fore.YELLOW + "Removing old files and folders...")
		shutil.rmtree("SOUNDS/" + str(args.language).split('-')[0])
		os.makedirs(filepath)
	elif not os.path.exists(filepath):
		os.makedirs(filepath)
		print(Fore.YELLOW + "Directories do not exist, creating filepath")
	else:
		print(Fore.YELLOW + "Filepath already exists")

def shorten_text(long_text, max_length):
	long_text = unicode(long_text)
	if(len(long_text) >= max_length):
		return long_text[:max_length-4] + "..."
	else:
		return long_text
	
	
def find_last_row(worksheet):
	lastRow = worksheet.max_row
	for i in range(2, worksheet.max_row):
		if (worksheet.cell(row=i, column=1).value == None) and (worksheet.cell(row=i, column=2).value == None) and (worksheet.cell(row=i, column=3).value == None):
			lastRow = i
			break
	return lastRow
		
def create_voice_from_list(filepath, worksheet):
	lastRow = find_last_row(worksheet)
	ignoredRows = worksheet.max_row - lastRow
	if ignoredRows <> 0:
		print(Fore.WHITE + "Ignored " + str(ignoredRows) + " Empty rows from file")
	else:
		lastRow += 1
	print("Generating " + str(lastRow - 2) + " Files")
	print(Back.GREEN + "\t\tFilename:\tDescription:\t\t\tText:")
	for r in tqdm(range(2, lastRow),file=sys.stdout):
		tqdm.write(u"\t\t{0:<10s}\t{1:<30s}\t{2:<40s}".format(shorten_text(worksheet.cell(row=r, column=1).value, 10),
			shorten_text(worksheet.cell(row=r, column=2).value, 30),shorten_text(worksheet.cell(row=r, column=3).value, 40)))
		synthesize_ssml('<speak>' + unicode(worksheet.cell(row=r, column=3).value) + '</speak>', filepath + unicode(worksheet.cell(row=r, column=1).value))

def create_voice_from_lines(start_index, end_index, filepath, worksheet):
	print(Back.GREEN + "\t\tFilename:\tDescription:\t\t\tText:")
	for r in tqdm(range(start_index, end_index),file=sys.stdout):
		tqdm.write(u"\t\t{0:<10s}\t{1:<30s}\t{2:<40s}".format(shorten_text(worksheet.cell(row=r, column=1).value, 10),
				shorten_text(worksheet.cell(row=r, column=2).value, 30),shorten_text(worksheet.cell(row=r, column=3).value, 40)))
		if not ((worksheet.cell(row=r, column=1).value == None) and (worksheet.cell(row=r, column=2).value == None) and (worksheet.cell(row=r, column=3).value == None)) and r >= 2:
			synthesize_ssml('<speak>' + unicode(worksheet.cell(row=r, column=3).value) + '</speak>', filepath + unicode(worksheet.cell(row=r, column=1).value))
		else:
			tqdm.write(Fore.YELLOW + "<-- Ignored")
		


if __name__ == '__main__':

	parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
	parser.add_argument('-f', '--file', required=True,
                       help='The name of the input file')
	parser.add_argument('-l', '--language', required=True,
                       help='If formated correctly this should be the language code. For example en-US for english')
	parser.add_argument('-v', '--voice', nargs='?',
                       help='Define the name of the TTS voice')
	parser.add_argument('-r', '--remove', action='store_true',
                       help='Use to clear all directories and start over fresh.')
	parser.add_argument('-o', '--override', action='store_true',
                       help='Use to enable overide if a file already exists. Use if you want to regenerate all voices from the excel file without affecting the folder structure. This will NOT delete any files!')
	parser.add_argument('-s', '--singleLine', type=int, nargs='?',
                       help='Use this option to regenerate one single line of the excel file. Add a minus to the line number to regenerate system voices, a positive value will regenerate user voices')
	parser.add_argument('-b', '--beginAtLine', type=int, nargs='?',
                       help='Use this to regenerate voices for multiple lines that you define. If you specify --beginAtLine you also have to specify --endAtLine. As with singleLine a negative number defines a line from system voices.')
	parser.add_argument('-e', '--endAtLine', type=int, nargs='?',
                       help='Use this to regenerate voices for multiple lines that you define. If you specify --endAtLine you also have to specify --beginAtLine. As with singleLine a negative number defines a line from system voices.')
	args = parser.parse_args()
	
	# Pre check arguments to make sure thet nothing is missconfigured.
	if (args.singleLine is not None and (args.beginAtLine is not None or args.endAtLine is not None)):
		print(Fore.RED + "You cannot try to gereate a sinle line and multiple lines at the same time. Please remove either the -s(--singleLine) argument or both of -b(--beginAtLine) AND -e(--endAtLine)")
		exit()
	if (args.beginAtLine is not None and args.endAtLine is None):
		print(Fore.RED + "The argument --beginAtLine requires --endAtLine")
		exit()
	elif (args.beginAtLine is None and args.endAtLine is not None):
		print(Fore.RED + "The argument --endAtLine requires --beginAtLine")
		exit()
	elif((args.beginAtLine < 0 and args.endAtLine < 0) and args.beginAtLine < args.endAtLine):
		print(Fore.RED + "Because you are generating system sounds --beginAtLine must be greater than --endAtLine")
		exit()
	elif(args.beginAtLine > args.endAtLine) and not (args.beginAtLine < 0 and args.endAtLine < 0):
		print(Fore.RED + "--endAtLine must be greater than --beginAtLine")
		exit()
	elif (args.beginAtLine is not None and args.endAtLine is not None):
		multiLine = 1
	else:
		multiLine = 0
		
	try:
		print("\n\n*************************************************************\n*** Starting up...\n*************************************************************\n")
		
		# Initialize the google cloud TTS API
		usingDefaultVoice = args.voice is None
		if usingDefaultVoice:
			usedVoice = str(args.language) + "-" + standardVoice
		else:
			usedVoice = str(args.language) + "-" + args.voice
		client = texttospeech.TextToSpeechClient()
		voice = texttospeech.types.VoiceSelectionParams(
			language_code = args.language,
			name = usedVoice)
		
		# Check for errors in the voice definition:
		voices = client.list_voices()
		voiceMatch = 0
		for oneVoice in voices.voices:
			#check if any voice mathes the request
			if(voice.name == oneVoice.name):
				voiceMatch = 1
				break
		
		if voiceMatch :
			print("Using '" + voice.name + "' for voice output")
		else:
			if usingDefaultVoice:
				# Try to find a voice that is supports the desired language:
				print(Fore.YELLOW + "The defined default voice does not exist. Using the first voice for your language in the list: ")
				selectedCountry_code = 0
				selectedName = 0
				for oneVoice in voices.voices:
					selectedCountry_code = 0
					for language_code in oneVoice.language_codes:
						if(language_code == str(args.language)):
							selectedCountry_code = language_code
							break
					if(selectedCountry_code is not 0):
						selectedName = oneVoice.name
						break
				if((selectedName is not 0) and (selectedCountry_code is not 0)):
					voice = texttospeech.types.VoiceSelectionParams(
						language_code = selectedCountry_code,
						name = selectedName)
					print(voice.name)
				else:
					print(Fore.RED + "Sorry, your Language is not supported by Google cloud TTS")
					exit()
			else:
				print(Fore.RED + "No voice found matching your request of: " + voice.name)
				print(Fore.WHITE + "Here is a list of all supported voices:")
				for oneVoice in voices.voices:
					# Display the voice's name and gender:
					ssml_voice_genders = ['SSML_VOICE_GENDER_UNSPECIFIED', 'MALE',
                              'FEMALE', 'NEUTRAL']
					print('Name: {} \t\tGender: {}'.format(oneVoice.name, ssml_voice_genders[oneVoice.ssml_gender]))
				exit()
		
		# Check if the input file exists:
		if not os.path.isfile(str(args.file)):
			print(Fore.RED + "This file does not exist: " + str(args.file))
			exit()
		
		# Try to determine which file type it is:
		try:
			inputFile = str(args.file)
			fileExention = inputFile.split('.')[len(inputFile.split('.')) - 1]
		except:
			print(Fore.RED + "Something is wrong with your filename")
			exit()
			
		# Check if file is a csv or a xlsx file. If you need checks for other file extentions add them here:
		if fileExention == 'csv':
			print(Fore.RED + "CSV files are not yet supported...")
			exit()
		elif fileExention == 'xlsx':
			wb = load_workbook(inputFile, data_only=True)
		else:
			print(Fore.RED + "Unsupported file")
			exit()
		
		
		# Start to process the request:
		filepathSystem = "SOUNDS/" + str(args.language).split('-')[0] + "/SYSTEM/"
		filepathUser = "SOUNDS/" + str(args.language).split('-')[0] + "/"
		create_filesystem(filepathSystem)
	
	
		if(args.singleLine is None and multiLine == 0):
			#generate system sounds
			print(Fore.CYAN + "Generating system voices")
			ws = wb[str(args.language) + '-system']
			create_voice_from_list(filepathSystem, ws)

			#generate user sounds
			print(Fore.CYAN + "Generating user defined voices")
			ws = wb[str(args.language) + '-user']
			create_voice_from_list(filepathUser, ws)
		elif(args.singleLine is not None):
			if(args.singleLine > 0):
				print(Fore.CYAN + "Generating user defined voice from line {}".format(str(args.singleLine)))
				ws = wb[str(args.language) + '-user']
				create_voice_from_lines(args.singleLine, args.singleLine + 1, filepathUser, ws)
			else:
				print(Fore.CYAN + "Generating system voice from line {}".format(str(-args.singleLine)))
				ws = wb[str(args.language) + '-system']
				create_voice_from_lines(-args.singleLine, (-args.singleLine) + 1, filepathSystem, ws)
		elif(multiLine == 1):
			if(args.beginAtLine > 0 and args.endAtLine > 0):
				print(Fore.CYAN + "Generating user defined voices from line {} to {}".format(str(args.beginAtLine), str(args.endAtLine)))
				ws = wb[str(args.language) + '-user']
				create_voice_from_lines(args.beginAtLine, args.endAtLine, filepathUser, ws)
			elif(args.beginAtLine < 0 and args.endAtLine < 0):
				print(Fore.CYAN + "Generating system voices from line {} to {}".format(str(-args.beginAtLine), str(-args.endAtLine)))
				ws = wb[str(args.language) + '-system']
				create_voice_from_lines(-args.beginAtLine, -args.endAtLine, filepathSystem, ws)
			else:
				ws = wb[str(args.language) + '-system']
				lastRow = find_last_row(ws)
				print(Fore.CYAN + "Generating system voices from line {} to {}".format(str(-args.beginAtLine), str(lastRow)))
				create_voice_from_lines(-args.beginAtLine, lastRow, filepathSystem, ws)
				
				print(Fore.CYAN + "Generating user defined voices from line 2 to {}".format(str(args.endAtLine)))
				ws = wb[str(args.language) + '-user']
				create_voice_from_lines(2, args.endAtLine + 1, filepathUser, ws)
				
				
		print("\n\n*************************************************************\n*** Done!!!\n*************************************************************\n")
	except KeyboardInterrupt:
		print(Fore.RED + "Exiting script")
		exit()